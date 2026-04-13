#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Gợi Ý Đặt Hàng
- File bán hàng : data.warehouse export (giao dịch thô, cột Date + Quantity)
- File tồn kho  : stock.quant export   (nhiều dòng/địa điểm, cột Số lượng)
"""

import tkinter as tk
from tkinter import filedialog, messagebox
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os, re, platform
from datetime import datetime
import sys
from PIL import Image, ImageTk

# Đường dẫn logo — dùng sys._MEIPASS khi chạy từ PyInstaller exe
if getattr(sys, 'frozen', False):
    _DIR = sys._MEIPASS
else:
    _DIR = os.path.dirname(os.path.abspath(__file__))
LOGO_PATH = os.path.join(_DIR, 'File_template', 'Bluecircle.png')

# ─── Font detection (SF Pro trên macOS, Helvetica Neue fallback) ──────────────
IS_MAC = platform.system() == 'Darwin'
SANS   = '.AppleSystemUIFont' if IS_MAC else 'Helvetica Neue'
MONO   = 'SF Mono'            if IS_MAC else 'Menlo'

# ─── Apple-style palette ─────────────────────────────────────────────────────
C = {
    'bg':         '#F2F2F7',   # iOS system background
    'card':       '#FFFFFF',
    'border':     '#C6C6C8',
    'sep':        '#E5E5EA',
    'primary':    '#007AFF',   # iOS blue
    'primary_dk': '#0062CC',
    'text':       '#000000',
    'text2':      '#8E8E93',
    'green':      '#34C759',
    'green_dk':   '#248A3D',
    'red_light':  '#FFF0F0',
    'red_text':   '#FF3B30',
    'blue_light': '#EAF3FF',
    'input_bg':   '#F2F2F7',
    'check_sel':  '#EAF3FF',
    'console_bg': '#1C1C1E',
    'console_fg': '#32D74B',
    'tag_bg':     '#EAF3FF',
    'tag_fg':     '#007AFF',
    'hdr_top':    '#1C1C1E',
}

SALES = {
    'sku':       'SKU',
    'name':      'Product Item',
    'brand':     'Brand',
    'category':  'Category',
    'date':      'Date',
    'qty':       'Quantity',
    'sale_team': 'Sale Team',
    'revenue':   'Revenue',
    'model':        'Model',
    'color':        'Color',
    'frame_size':   'Frame Size',
    'sub_category': 'Sub Category',
}
INV = {
    'sku':          'Sản phẩm/Mã nội bộ',
    'qty':          'Số lượng',
    'qty_reserved': 'Số lượng bảo lưu',
    'brand':        'Sản phẩm/Brand/Display Name',
    'category':     'Sản phẩm/Nhóm Điểm bán lẻ/Tên hiển thị',
}


def strip_sku_prefix(t: str) -> str:
    return re.sub(r'^\[.*?\]\s*', '', str(t)).strip()


# ─── Reusable widgets ─────────────────────────────────────────────────────────

def make_card(parent, title, subtitle=''):
    """Apple-style card: white frame with hairline border + separator."""
    outer = tk.Frame(parent, bg=C['border'], padx=1, pady=1)
    inner = tk.Frame(outer, bg=C['card'])
    inner.pack(fill=tk.BOTH, expand=True)

    hdr = tk.Frame(inner, bg=C['card'], padx=20, pady=12)
    hdr.pack(fill=tk.X)
    tk.Label(hdr, text=title, font=(SANS, 13, 'bold'),
             bg=C['card'], fg=C['text']).pack(side=tk.LEFT)
    if subtitle:
        tk.Label(hdr, text=subtitle, font=(SANS, 11),
                 bg=C['card'], fg=C['text2']).pack(side=tk.LEFT, padx=(8, 0))

    tk.Frame(inner, bg=C['sep'], height=1).pack(fill=tk.X)

    body = tk.Frame(inner, bg=C['card'], padx=20, pady=14)
    body.pack(fill=tk.BOTH, expand=True)

    return outer, body


def make_btn(parent, text, command, style='primary', small=False):
    """Flat Apple-style button."""
    if style == 'primary':
        bg, fg, abg = C['primary'],    'white', C['primary_dk']
    elif style == 'green':
        bg, fg, abg = C['green'],      'white', C['green_dk']
    elif style == 'ghost':
        bg, fg, abg = C['blue_light'], C['primary'], '#D0E8FF'
    else:
        bg, fg, abg = C['sep'],        C['text'], '#D0D0D5'

    font = (SANS, 10 if small else 13, 'bold')
    px, py = (12, 5) if small else (22, 10)
    return tk.Button(parent, text=text, command=command,
                     font=font, bg=bg, fg=fg, bd=0, relief='flat',
                     cursor='hand2', padx=px, pady=py,
                     activebackground=abg, activeforeground=fg)


# ─── Loading popup với spinner ───────────────────────────────────────────────

class LoadingPopup:
    """
    Popup nhỏ giữa màn hình, có arc xoay tròn và message tuỳ chỉnh.
    Dùng:  popup = LoadingPopup(root, 'message')  →  popup.close()
    """
    _SIZE   = 56   # đường kính vòng spinner
    _THICK  = 5
    _GAP    = 270  # phần trống của arc (°)
    _SPEED  = 18   # ms giữa mỗi frame

    def __init__(self, parent: tk.Tk, message: str):
        self._running = False
        self._angle   = 0

        # ── Cửa sổ không có title bar ─────────────────────────────────────
        self.top = tk.Toplevel(parent)
        self.top.overrideredirect(True)
        self.top.configure(bg=C['card'])
        self.top.attributes('-topmost', True)
        self.top.resizable(False, False)

        # viền mỏng quanh popup
        border = tk.Frame(self.top, bg=C['border'], padx=1, pady=1)
        border.pack(fill=tk.BOTH, expand=True)
        inner  = tk.Frame(border, bg=C['card'], padx=36, pady=28)
        inner.pack(fill=tk.BOTH, expand=True)

        # spinner canvas
        s = self._SIZE
        self._cv = tk.Canvas(inner, width=s, height=s,
                             bg=C['card'], highlightthickness=0)
        self._cv.pack()

        # vòng nền mờ
        m = self._THICK // 2 + 1
        self._cv.create_oval(m, m, s - m, s - m,
                             outline=C['sep'], width=self._THICK)
        # arc xoay
        self._arc_id = self._cv.create_arc(
            m, m, s - m, s - m,
            start=0, extent=360 - self._GAP,
            outline=C['primary'], width=self._THICK, style='arc'
        )

        # dòng message (hỗ trợ xuống dòng)
        tk.Label(inner, text=message,
                 font=(SANS, 12), bg=C['card'], fg=C['text'],
                 justify='center', wraplength=280,
                 pady=(14)).pack(pady=(14, 0))

        # căn giữa so với parent
        self.top.update_idletasks()
        pw = parent.winfo_width()
        ph = parent.winfo_height()
        px = parent.winfo_rootx()
        py = parent.winfo_rooty()
        ww = self.top.winfo_reqwidth()
        wh = self.top.winfo_reqheight()
        x  = px + (pw - ww) // 2
        y  = py + (ph - wh) // 2
        self.top.geometry(f'+{x}+{y}')

        # chặn tương tác với cửa sổ chính
        self.top.grab_set()

        # bắt đầu animate
        self._running = True
        self._animate()

    def _animate(self):
        if not self._running:
            return
        self._angle = (self._angle - 8) % 360   # quay ngược chiều kim đồng hồ
        self._cv.itemconfigure(self._arc_id, start=self._angle)
        self._cv.after(self._SPEED, self._animate)

    def close(self):
        self._running = False
        try:
            self.top.grab_release()
            self.top.destroy()
        except tk.TclError:
            pass


# ─── Scrollable multi-select panel ───────────────────────────────────────────

class FilterPanel(tk.Frame):
    """
    Scrollable checkbox panel với ô tìm kiếm.
    Không chọn gì = không lọc (lấy tất cả).
    """
    def __init__(self, parent, **kw):
        super().__init__(parent, bg=C['card'], **kw)
        self._vars: dict[str, tk.BooleanVar] = {}
        self._all_items: list = []
        self._widgets: list = []
        self._placeholder_active = True
        self._build()

    # ── scroll helper ─────────────────────────────────────────────────────────
    def _on_scroll(self, event):
        """Route MouseWheel to this panel's canvas, stop propagation."""
        if event.delta:
            # Windows / macOS: delta is ±120 per notch (or smaller for trackpad)
            units = int(-1 * (event.delta / 120)) or (-1 if event.delta > 0 else 1)
        elif event.num == 4:
            units = -1
        else:
            units = 1
        self._canvas.yview_scroll(units, 'units')
        return 'break'   # ← stops the event from reaching the outer canvas

    def _build(self):
        # ── Ô tìm kiếm ───────────────────────────────────────────────────────
        search_wrap = tk.Frame(self, bg=C['border'], padx=1, pady=1)
        search_wrap.pack(fill=tk.X, pady=(0, 8))
        search_inner = tk.Frame(search_wrap, bg=C['input_bg'])
        search_inner.pack(fill=tk.X)
        tk.Label(search_inner, text='⌕', font=(SANS, 14),
                 bg=C['input_bg'], fg=C['text2']).pack(side=tk.LEFT, padx=(10, 4))
        self._search_var = tk.StringVar()
        self._search_entry = tk.Entry(
            search_inner,
            textvariable=self._search_var,
            font=(SANS, 12), bg=C['input_bg'], fg=C['text2'],
            relief='flat', bd=0, insertbackground=C['primary'])
        self._search_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, ipady=7, padx=(0, 10))
        self._search_entry.insert(0, 'Tìm kiếm...')
        self._search_entry.bind('<FocusIn>',  self._on_search_focus_in)
        self._search_entry.bind('<FocusOut>', self._on_search_focus_out)
        self._search_var.trace_add('write', lambda *_: self._apply_search())

        # ── Toolbar ───────────────────────────────────────────────────────────
        tb = tk.Frame(self, bg=C['card'])
        tb.pack(fill=tk.X, pady=(0, 8))
        make_btn(tb, '✓ Tất Cả',  lambda: self._set_all(True),  style='ghost',   small=True).pack(side=tk.LEFT, padx=(0, 6))
        make_btn(tb, '✕ Bỏ Chọn', lambda: self._set_all(False), style='neutral', small=True).pack(side=tk.LEFT)
        self.lbl_count = tk.Label(tb, text='—  chưa tải', font=(SANS, 10),
                                   bg=C['card'], fg=C['text2'])
        self.lbl_count.pack(side=tk.RIGHT)

        # ── Scroll area ───────────────────────────────────────────────────────
        wrap = tk.Frame(self, bg=C['border'], padx=1, pady=1)
        wrap.pack(fill=tk.BOTH, expand=True)
        inner_bg = tk.Frame(wrap, bg=C['card'])
        inner_bg.pack(fill=tk.BOTH, expand=True)

        self._canvas = tk.Canvas(inner_bg, bg=C['card'], height=160,
                                  highlightthickness=0, bd=0)
        vsb = tk.Scrollbar(inner_bg, orient='vertical', command=self._canvas.yview)
        self._canvas.configure(yscrollcommand=vsb.set)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)
        self._canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        self._inner = tk.Frame(self._canvas, bg=C['card'])
        self._win   = self._canvas.create_window((0, 0), window=self._inner, anchor='nw')

        self._inner.bind('<Configure>', lambda e: self._canvas.configure(
            scrollregion=self._canvas.bbox('all')))
        self._canvas.bind('<Configure>', lambda e: self._canvas.itemconfig(
            self._win, width=e.width))

        # Scroll bindings – both canvas AND inner frame, return 'break' to stop bubbling
        self._canvas.bind('<MouseWheel>', self._on_scroll)
        self._inner.bind('<MouseWheel>',  self._on_scroll)
        self._canvas.bind('<Button-4>',   self._on_scroll)   # Linux
        self._canvas.bind('<Button-5>',   self._on_scroll)

        tk.Label(self._inner, text='Nhấn  "Đọc Files"  để tải danh sách',
                 font=(SANS, 10, 'italic'),
                 bg=C['card'], fg=C['text2'], pady=28).pack()

    # ── Search ────────────────────────────────────────────────────────────────

    def _on_search_focus_in(self, e):
        if self._placeholder_active:
            self._search_entry.delete(0, tk.END)
            self._search_entry.config(fg=C['text'])
            self._placeholder_active = False

    def _on_search_focus_out(self, e):
        if not self._search_entry.get().strip():
            self._search_entry.delete(0, tk.END)
            self._search_entry.insert(0, 'Tìm kiếm...')
            self._search_entry.config(fg=C['text2'])
            self._placeholder_active = True

    def _apply_search(self):
        if self._placeholder_active:
            self._render(self._all_items)
            return
        keyword = self._search_var.get().strip().lower()
        filtered = [i for i in self._all_items if keyword in i.lower()] if keyword else self._all_items
        self._render(filtered)

    def _render(self, items: list):
        """Vẽ lại checkbox list theo danh sách items (giữ nguyên trạng thái tick)."""
        for w in self._inner.winfo_children():
            w.destroy()
        self._widgets.clear()

        if not items:
            tk.Label(self._inner, text='Không tìm thấy kết quả',
                     font=(SANS, 10, 'italic'),
                     bg=C['card'], fg=C['text2'], pady=20).pack()
            return

        cols = 2
        for idx, item in enumerate(items):
            if item not in self._vars:
                self._vars[item] = tk.BooleanVar(value=False)
            cb = tk.Checkbutton(
                self._inner, text=item, variable=self._vars[item],
                font=(SANS, 11), bg=C['card'], fg=C['text'],
                selectcolor=C['check_sel'], activebackground=C['card'],
                anchor='w', command=self._update_count,
            )
            r, c = divmod(idx, cols)
            cb.grid(row=r, column=c, sticky='w', padx=10, pady=3)
            # Scroll trên checkbox cũng cuộn panel, không bắn lên canvas ngoài
            cb.bind('<MouseWheel>', self._on_scroll)
            cb.bind('<Button-4>',   self._on_scroll)
            cb.bind('<Button-5>',   self._on_scroll)
            self._widgets.append(cb)

        for c in range(cols):
            self._inner.columnconfigure(c, weight=1)

    def populate(self, items: list):
        self._all_items = items
        self._vars.clear()
        self._widgets.clear()
        # Reset ô tìm kiếm về placeholder
        self._search_entry.delete(0, tk.END)
        self._search_entry.insert(0, 'Tìm kiếm...')
        self._search_entry.config(fg=C['text2'])
        self._placeholder_active = True

        self._render(items)
        self._update_count()

    def _set_all(self, val: bool):
        # Chỉ áp dụng cho items đang hiển thị (sau filter tìm kiếm)
        if self._placeholder_active:
            visible = self._all_items
        else:
            keyword = self._search_entry.get().lower().strip()
            visible = [i for i in self._all_items if keyword in i.lower()] if keyword else self._all_items
        for item in visible:
            if item in self._vars:
                self._vars[item].set(val)
        self._update_count()

    def _update_count(self):
        total    = len(self._vars)
        selected = sum(1 for v in self._vars.values() if v.get())
        if total == 0:
            self.lbl_count.config(text='—  chưa tải', fg=C['text2'])
        elif selected == 0:
            self.lbl_count.config(text='Trống = lấy tất cả', fg=C['text2'])
        else:
            self.lbl_count.config(text=f'{selected} / {total} đã chọn',
                                   fg=C['primary'])

    def selected(self) -> list:
        """Trả về list item đã chọn. Rỗng = không lọc (lấy tất cả)."""
        return [k for k, v in self._vars.items() if v.get()]

    def all_items(self) -> list:
        return list(self._vars.keys())


# ─── Main App ─────────────────────────────────────────────────────────────────

class App:
    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title('Gợi Ý Đặt Hàng')
        self.root.geometry('1180x740')
        self.root.configure(bg=C['bg'])
        self.root.resizable(True, True)

        self.v_sales    = tk.StringVar()
        self.v_inv      = tk.StringVar()
        self.v_months   = tk.StringVar(value='')   # bắt buộc người dùng nhập
        self.v_leadtime = tk.StringVar(value='')   # bắt buộc người dùng nhập
        # không cần chọn kênh — luôn tính cả BL lẫn BS

        self._build_ui()
        self.log('✅ Ứng dụng khởi động thành công.')
        self.log('📌 Bước 1: Chọn 2 file  →  Bước 2: Đọc Files  →  Bước 3: Lọc (tuỳ chọn)  →  Bước 4: Nhập Tồn tối thiểu + Leadtime  →  Bước 5: Tạo Gợi Ý.')

    # ── Build UI ─────────────────────────────────────────────────────────────

    def _build_ui(self):
        # ── Header ───────────────────────────────────────────────────────────
        HDR_BG = '#FFFFFF'
        hdr = tk.Frame(self.root, bg=HDR_BG)
        hdr.pack(fill=tk.X)

        hdr_inner = tk.Frame(hdr, bg=HDR_BG, padx=32, pady=18)
        hdr_inner.pack(fill=tk.X)

        # Logo + tiêu đề cùng hàng ngang
        title_row = tk.Frame(hdr_inner, bg=HDR_BG)
        title_row.pack(anchor='w')

        # Logo bên trái tiêu đề
        try:
            pil_img = Image.open(LOGO_PATH).convert('RGBA')
            logo_h  = 48
            logo_w  = int(pil_img.width * logo_h / pil_img.height)
            pil_img = pil_img.resize((logo_w, logo_h), Image.LANCZOS)
            bg_img  = Image.new('RGBA', (logo_w, logo_h), (255, 255, 255, 255))
            bg_img.paste(pil_img, mask=pil_img.split()[3])
            self._logo_img = ImageTk.PhotoImage(bg_img.convert('RGB'))
            tk.Label(title_row, image=self._logo_img,
                     bg=HDR_BG).pack(side=tk.LEFT, anchor='center', padx=(0, 14))
        except Exception:
            pass

        tk.Label(title_row, text='Gợi Ý Đặt Hàng',
                 font=(SANS, 24, 'bold'),
                 bg=HDR_BG, fg=C['primary']).pack(side=tk.LEFT, anchor='center')

        tk.Label(hdr_inner,
                 text='Phân tích dữ liệu bán hàng & tồn kho  ·  Xuất file Excel gợi ý đặt hàng',
                 font=(SANS, 12), bg=HDR_BG, fg='#98989F').pack(anchor='w', pady=(6, 0))

        # Đường kẻ phân cách dưới header
        tk.Frame(self.root, bg='#E5E5EA', height=1).pack(fill=tk.X)

        # ── Body (fixed layout, no outer scroll) ─────────────────────────────
        self.body = tk.Frame(self.root, bg=C['bg'])
        self.body.pack(fill=tk.BOTH, expand=True)

        self._build_body()

    def _build_body(self):
        b = self.body
        b.columnconfigure(0, weight=2)
        b.columnconfigure(1, weight=1)
        b.rowconfigure(0, weight=1)

        # ── Cột trái: scrollable (fix màn hình nhỏ Windows) ──────────────────
        left_outer = tk.Frame(b, bg=C['bg'])
        left_outer.grid(row=0, column=0, sticky='nsew', padx=(22, 8), pady=18)
        left_outer.columnconfigure(0, weight=1)
        left_outer.rowconfigure(0, weight=1)

        self._left_canvas = tk.Canvas(left_outer, bg=C['bg'], highlightthickness=0)
        left_vsb = tk.Scrollbar(left_outer, orient='vertical',
                                command=self._left_canvas.yview)
        self._left_canvas.configure(yscrollcommand=left_vsb.set)
        left_vsb.pack(side=tk.RIGHT, fill=tk.Y)
        self._left_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        left = tk.Frame(self._left_canvas, bg=C['bg'])
        _lwin = self._left_canvas.create_window((0, 0), window=left, anchor='nw')
        left.bind('<Configure>', lambda e: self._left_canvas.configure(
            scrollregion=self._left_canvas.bbox('all')))
        self._left_canvas.bind('<Configure>', lambda e: self._left_canvas.itemconfig(
            _lwin, width=e.width))

        def _left_scroll(event):
            if event.delta:
                units = int(-1 * (event.delta / 120)) or (-1 if event.delta > 0 else 1)
            elif event.num == 4:
                units = -1
            else:
                units = 1
            self._left_canvas.yview_scroll(units, 'units')
            return 'break'

        self._left_canvas.bind('<MouseWheel>', _left_scroll)
        self._left_canvas.bind('<Button-4>',   _left_scroll)
        self._left_canvas.bind('<Button-5>',   _left_scroll)
        left.bind('<MouseWheel>', _left_scroll)
        left.bind('<Button-4>',   _left_scroll)
        left.bind('<Button-5>',   _left_scroll)

        left.columnconfigure(0, weight=1)

        # ── 1. Import Files ───────────────────────────────────────────────────
        card, body = make_card(left, '① Import Files')
        self._file_row(body, 'File Bán Hàng', self.v_sales)
        tk.Frame(body, bg=C['sep'], height=1).pack(fill=tk.X, pady=6)
        self._file_row(body, 'File Tồn Kho', self.v_inv)
        btn_row = tk.Frame(body, bg=C['card'])
        btn_row.pack(fill=tk.X, pady=(12, 0))
        self.btn_read = make_btn(btn_row, '  Đọc Files & Tải Bộ Lọc  →',
                                  self._read_files, style='green')
        self.btn_read.pack(side=tk.RIGHT)
        card.pack(fill=tk.X, pady=(0, 10))

        # ── 2. Bộ lọc Brand + Category ───────────────────────────────────────
        filter_row = tk.Frame(left, bg=C['bg'])
        filter_row.pack(fill=tk.X, pady=(0, 10))
        filter_row.columnconfigure(0, weight=1)
        filter_row.columnconfigure(1, weight=1)

        bc, bbody = make_card(filter_row, '② Thương Hiệu', '(tuỳ chọn · trống = tất cả)')
        self.brand_panel = FilterPanel(bbody)
        self.brand_panel.pack(fill=tk.BOTH, expand=True)
        bc.grid(row=0, column=0, sticky='nsew', padx=(0, 6))

        cc, cbody = make_card(filter_row, '③ Danh Mục', '(tuỳ chọn · trống = tất cả)')
        self.cat_panel = FilterPanel(cbody)
        self.cat_panel.pack(fill=tk.BOTH, expand=True)
        cc.grid(row=0, column=1, sticky='nsew', padx=(6, 0))

        filter_row.bind('<MouseWheel>', _left_scroll)
        filter_row.bind('<Button-4>',   _left_scroll)
        filter_row.bind('<Button-5>',   _left_scroll)

        # ── 3. Phân kênh Sale Team (BL / BS) ─────────────────────────────────
        chan_card, chan_body = make_card(left, '④ Phân Kênh Sale Team',
                                         '(tuỳ chọn · trống = dùng tất cả team)')
        chan_body.config(pady=10)
        chan_card.pack(fill=tk.X, pady=(0, 10))

        tk.Label(chan_body,
                 text='Sau khi Đọc Files, chọn Sale Team thuộc từng kênh. Trống = dùng tất cả team cho kênh đó.',
                 font=(SANS, 10), bg=C['card'], fg=C['text2'],
                 anchor='w', wraplength=460, justify='left').pack(fill=tk.X, pady=(0, 10))

        chan_cols_frame = tk.Frame(chan_body, bg=C['card'])
        chan_cols_frame.pack(fill=tk.X)
        chan_cols_frame.columnconfigure(0, weight=1)
        chan_cols_frame.columnconfigure(1, weight=1)

        # ─ BL ────────────────────────────────────────────────────────────────
        bl_col = tk.Frame(chan_cols_frame, bg=C['card'])
        bl_col.grid(row=0, column=0, sticky='nsew', padx=(0, 6))
        tk.Label(bl_col, text='Bán Lẻ (Retail)',
                 font=(SANS, 12, 'bold'), bg=C['card'], fg=C['primary'],
                 anchor='w').pack(fill=tk.X, pady=(0, 4))
        self.bl_team_panel = FilterPanel(bl_col)
        self.bl_team_panel.pack(fill=tk.BOTH, expand=True)

        # ─ BS ────────────────────────────────────────────────────────────────
        bs_col = tk.Frame(chan_cols_frame, bg=C['card'])
        bs_col.grid(row=0, column=1, sticky='nsew', padx=(6, 0))
        tk.Label(bs_col, text='Bán Sỉ (Wholesale)',
                 font=(SANS, 12, 'bold'), bg=C['card'], fg='#6D28D9',
                 anchor='w').pack(fill=tk.X, pady=(0, 4))
        self.bs_team_panel = FilterPanel(bs_col)
        self.bs_team_panel.pack(fill=tk.BOTH, expand=True)

        # ── 4. Tồn kho tối thiểu + Leadtime (bắt buộc nhập, không có mặc định)
        settings_row = tk.Frame(left, bg=C['bg'])
        settings_row.pack(fill=tk.X, pady=(0, 10))
        settings_row.columnconfigure(0, weight=1)
        settings_row.columnconfigure(1, weight=1)

        mc, mbody = make_card(settings_row, '⑤ Tồn Kho Tối Thiểu', '(bắt buộc nhập)')
        mrow = tk.Frame(mbody, bg=C['card'])
        mrow.pack(fill=tk.X)
        tk.Label(mrow, text='Số tháng :', font=(SANS, 13),
                 bg=C['card'], fg=C['text']).pack(side=tk.LEFT)
        tk.Entry(mrow, textvariable=self.v_months, width=6,
                 font=(SANS, 13), bd=1, relief='solid',
                 bg=C['input_bg'], fg=C['text'],
                 insertbackground=C['primary']).pack(side=tk.LEFT, padx=(10, 0), ipady=4)
        tk.Label(mrow, text='tháng', font=(SANS, 11),
                 bg=C['card'], fg=C['text2']).pack(side=tk.LEFT, padx=(6, 0))
        mc.grid(row=0, column=0, sticky='nsew', padx=(0, 6))

        lc2, lbody2 = make_card(settings_row, '⑥ Leadtime Về Hàng', '(bắt buộc nhập)')
        lrow = tk.Frame(lbody2, bg=C['card'])
        lrow.pack(fill=tk.X)
        tk.Label(lrow, text='Số tháng :', font=(SANS, 13),
                 bg=C['card'], fg=C['text']).pack(side=tk.LEFT)
        tk.Entry(lrow, textvariable=self.v_leadtime, width=6,
                 font=(SANS, 13), bd=1, relief='solid',
                 bg=C['input_bg'], fg=C['text'],
                 insertbackground=C['primary']).pack(side=tk.LEFT, padx=(10, 0), ipady=4)
        tk.Label(lrow, text='tháng', font=(SANS, 11),
                 bg=C['card'], fg=C['text2']).pack(side=tk.LEFT, padx=(6, 0))
        lc2.grid(row=0, column=1, sticky='nsew', padx=(6, 0))

        # ── CTA button ────────────────────────────────────────────────────────
        self.btn_run = make_btn(left, '  🚀   Tạo Gợi Ý Đặt Hàng  ',
                                 self.run, style='primary')
        self.btn_run.config(font=(SANS, 15, 'bold'), pady=15)
        self.btn_run.pack(fill=tk.X)

        # ── Cột phải: log (không cần scroll) ─────────────────────────────────
        right = tk.Frame(b, bg=C['bg'])
        right.grid(row=0, column=1, sticky='nsew', padx=(8, 22), pady=18)
        right.columnconfigure(0, weight=1)
        right.rowconfigure(0, weight=1)

        lc, lbody = make_card(right, 'Nhật Ký')
        lbody.config(padx=0, pady=0)
        self.log_box = tk.Text(lbody, font=(MONO, 10),
                                bg=C['console_bg'], fg=C['console_fg'],
                                padx=16, pady=10, wrap=tk.WORD, bd=0,
                                insertbackground=C['console_fg'])
        self.log_box.pack(fill=tk.BOTH, expand=True)
        lc.grid(row=0, column=0, sticky='nsew')

    # ── Widgets helpers ───────────────────────────────────────────────────────

    def _file_row(self, parent, label, var):
        f = tk.Frame(parent, bg=C['card'])
        f.pack(fill=tk.X)

        tk.Label(f, text=label, font=(SANS, 13),
                 bg=C['card'], fg=C['text'], width=16, anchor='w').pack(side=tk.LEFT)

        entry = tk.Entry(f, textvariable=var, font=(SANS, 11),
                         bg=C['input_bg'], fg=C['text2'], relief='flat',
                         bd=0, state='readonly', readonlybackground=C['input_bg'])
        entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(8, 10), ipady=6)

        make_btn(f, 'Chọn File', lambda v=var: self._browse(v),
                 style='ghost', small=True).pack(side=tk.RIGHT)

    # ── Helpers ───────────────────────────────────────────────────────────────

    def _browse(self, var):
        path = filedialog.askopenfilename(
            filetypes=[('Excel files', '*.xlsx *.xls'), ('All files', '*.*')])
        if path:
            var.set(path)
            self.log(f'📂  {os.path.basename(path)}')

    def log(self, msg: str):
        ts = datetime.now().strftime('%H:%M:%S')
        self.log_box.insert(tk.END, f'[{ts}]  {msg}\n')
        self.log_box.see(tk.END)
        self.root.update()

    # ── Đọc files & populate filters ─────────────────────────────────────────

    def _read_files(self):
        s = self.v_sales.get()
        i = self.v_inv.get()
        if not s:
            messagebox.showerror('Lỗi', 'Vui lòng chọn file bán hàng!'); return
        if not i:
            messagebox.showerror('Lỗi', 'Vui lòng chọn file tồn kho!'); return

        popup = LoadingPopup(
            self.root,
            'Sếp đợi chút\nem đang tính toán,\nem biết sếp tính lóng như kem 🍦'
        )
        try:
            self.btn_read.config(state='disabled', text='⏳  Đang đọc…')
            self.log('─── Đọc files ───')

            sales = pd.read_excel(s)
            sales.columns = [str(c).strip() for c in sales.columns]
            self._check_cols(sales, [SALES['brand'], SALES['category']], 'file bán hàng')

            # Lọc SERVICE/COUPON
            sales = sales[
                sales[SALES['sku']].notna() &
                (~sales[SALES['sku']].astype(str).str.upper().str.contains('COUPON|DISCOUNT', na=False)) &
                (sales[SALES['category']].astype(str).str.upper() != 'SERVICE')
            ]

            inv = pd.read_excel(i)
            inv.columns = [str(c).strip() for c in inv.columns]

            def clean(series):
                return set(series.dropna().astype(str).str.strip()
                           .replace('', pd.NA).dropna().unique())

            # Brand
            brands = sorted(
                clean(sales[SALES['brand']]) |
                (clean(inv[INV['brand']]) if INV['brand'] in inv.columns else set()),
                key=str.upper)

            # Category
            cats = sorted(
                clean(sales[SALES['category']]) |
                (clean(inv[INV['category']]) if INV['category'] in inv.columns else set()),
                key=str.upper)

            self.brand_panel.populate(brands)
            self.cat_panel.populate(cats)

            # Sale Teams cho phân kênh BL / BS
            teams: list = []
            if SALES['sale_team'] in sales.columns:
                teams = sorted(clean(sales[SALES['sale_team']]), key=str.upper)
            self.bl_team_panel.populate(teams)
            self.bs_team_panel.populate(teams)

            self.log(f'✅  {len(brands)} thương hiệu  ·  {len(cats)} danh mục  ·  {len(teams)} sale team  →  sẵn sàng lọc')

        except Exception as e:
            self.log(f'❌  {e}')
            messagebox.showerror('Lỗi đọc file', str(e))
        finally:
            popup.close()
            self.btn_read.config(state='normal', text='  Đọc Files & Tải Bộ Lọc  →')

    # ── Run ───────────────────────────────────────────────────────────────────

    def run(self):
        s = self.v_sales.get()
        i = self.v_inv.get()

        if not s: messagebox.showerror('Lỗi', 'Vui lòng chọn file bán hàng!'); return
        if not i: messagebox.showerror('Lỗi', 'Vui lòng chọn file tồn kho!'); return

        # Validate tồn kho tối thiểu + leadtime (bắt buộc nhập)
        m_str  = self.v_months.get().strip()
        lt_str = self.v_leadtime.get().strip()
        if not m_str:
            messagebox.showerror('Lỗi',
                'Vui lòng nhập Tồn Kho Tối Thiểu (số tháng)!\n(Mục ⑤)')
            return
        if not lt_str:
            messagebox.showerror('Lỗi',
                'Vui lòng nhập Leadtime Về Hàng (số tháng)!\n(Mục ⑥)')
            return
        try:
            m  = max(1, int(m_str))
            lt = max(0, int(lt_str))
        except ValueError:
            messagebox.showerror('Lỗi', 'Tồn kho tối thiểu và leadtime phải là số nguyên!'); return

        sel_brands = self.brand_panel.selected()
        sel_cats   = self.cat_panel.selected()
        bl_teams   = self.bl_team_panel.selected()
        bs_teams   = self.bs_team_panel.selected()

        brand_desc = ', '.join(sel_brands) if sel_brands else 'Tất cả'
        cat_desc   = ', '.join(sel_cats)   if sel_cats   else 'Tất cả'

        popup = LoadingPopup(
            self.root,
            'Sếp đợi em chút nha,\nem đang xử lý đây 🚀'
        )
        try:
            self.btn_run.config(state='disabled', text='⏳   Đang xử lý…')
            self.log(f'─── Tồn min: {m}T · Leadtime: {lt}T · Brand: {brand_desc} · Danh mục: {cat_desc} ───')
            df  = self._process(s, i, m, lt, sel_brands, sel_cats,
                                bl_teams=bl_teams, bs_teams=bs_teams)
            out = self._export(df, m, lt, sel_brands, sel_cats)
            self.log(f'✅  Hoàn thành → {out}')
            messagebox.showinfo('Thành công', f'Đã tạo file gợi ý đặt hàng!\n\n📄 {out}')
            if sys.platform == 'win32':
                os.startfile(os.path.dirname(out))
        except Exception as e:
            self.log(f'❌  {e}')
            messagebox.showerror('Lỗi xử lý', str(e))
        finally:
            popup.close()
            self.btn_run.config(state='normal', text='  🚀   Tạo Gợi Ý Đặt Hàng  ')

    # ── Core logic ────────────────────────────────────────────────────────────

    def _process(self, sales_path, inv_path, months, leadtime, sel_brands, sel_cats,
                 bl_teams=None, bs_teams=None) -> pd.DataFrame:

        # 1. Đọc & làm sạch bán hàng
        self.log('📊  Đọc file bán hàng…')
        sales = pd.read_excel(sales_path)
        sales.columns = [str(c).strip() for c in sales.columns]
        self._check_cols(sales, list(SALES.values()), 'file bán hàng')

        sales[SALES['date']] = pd.to_datetime(sales[SALES['date']], errors='coerce')
        sales = sales.dropna(subset=[SALES['date']])
        sales['_month'] = sales[SALES['date']].dt.to_period('M')
        sales[SALES['qty']] = pd.to_numeric(sales[SALES['qty']], errors='coerce').fillna(0)

        before = len(sales)
        sales = sales[
            sales[SALES['sku']].notna() &
            (sales[SALES['sku']].astype(str).str.strip() != '') &
            (~sales[SALES['sku']].astype(str).str.upper().str.contains('COUPON|DISCOUNT', na=False)) &
            (sales[SALES['category']].astype(str).str.upper() != 'SERVICE')
        ]
        if len(sales) < before:
            self.log(f'   ⚠️  Lọc {before - len(sales)} dòng không hợp lệ')

        # 2. Áp dụng bộ lọc Brand + Category (trống = lấy tất cả)
        if sel_brands:
            sales = sales[sales[SALES['brand']].astype(str).str.strip().isin(sel_brands)]
        if sel_cats:
            sales = sales[sales[SALES['category']].astype(str).str.strip().isin(sel_cats)]

        if sales.empty:
            raise ValueError('Không có dữ liệu sau khi lọc. Hãy thử chọn lại Brand / Danh mục.')

        self.log(f'   {len(sales):,} giao dịch  ·  {sales["_month"].nunique()} tháng  ·  {sales[SALES["sku"]].nunique():,} SKU')

        # 3. Thông tin sản phẩm
        has_model      = SALES['model']        in sales.columns
        has_color      = SALES['color']        in sales.columns
        has_frame_size = SALES['frame_size']   in sales.columns
        has_sub_cat    = SALES['sub_category'] in sales.columns
        extra_cols  = (
            ([SALES['model']]        if has_model      else []) +
            ([SALES['color']]        if has_color      else []) +
            ([SALES['frame_size']]   if has_frame_size else []) +
            ([SALES['sub_category']] if has_sub_cat    else [])
        )
        extra_names = (
            (['Model']        if has_model      else []) +
            (['Color']        if has_color      else []) +
            (['Frame Size']   if has_frame_size else []) +
            (['Sub Category'] if has_sub_cat    else [])
        )
        prod_info   = (
            sales[[SALES['sku'], SALES['name'], SALES['brand'], SALES['category']] + extra_cols]
            .drop_duplicates(subset=[SALES['sku']], keep='first').copy()
        )
        prod_info.columns = ['SKU', 'Tên Sản Phẩm', 'Brand', 'Category'] + extra_names
        prod_info['Tên Sản Phẩm'] = prod_info['Tên Sản Phẩm'].apply(strip_sku_prefix)
        prod_info['SKU'] = prod_info['SKU'].astype(str).str.strip()

        # 4. Pivot theo tháng (max 12 tháng gần nhất)
        monthly = (sales.groupby([SALES['sku'], '_month'])[SALES['qty']]
                   .sum().reset_index())
        monthly.columns = ['SKU', 'Month', 'Qty']
        monthly['SKU'] = monthly['SKU'].astype(str).str.strip()

        all_months = sorted(monthly['Month'].unique())
        use_months = all_months   # lấy tất cả tháng có trong file
        monthly    = monthly[monthly['Month'].isin(use_months)]

        pivot = monthly.pivot_table(index='SKU', columns='Month',
                                    values='Qty', fill_value=0).reset_index()
        mcmap = {p: f'Th.{p.month}/{p.year}' for p in use_months}
        pivot.rename(columns=mcmap, inplace=True)
        mlabels = [mcmap[p] for p in use_months]
        self.log(f'   Tháng: {mlabels[0]}  →  {mlabels[-1]}')

        # 5. Ghép + tính toán
        result = prod_info.merge(pivot, on='SKU', how='left')
        for ml in mlabels:
            if ml not in result.columns:
                result[ml] = 0

        md = result[mlabels].fillna(0)

        # Nhóm tháng theo năm → Tổng YYYY + TB YYYY
        from collections import defaultdict
        year_months_map = defaultdict(list)
        for ml in mlabels:
            year = ml.split('/')[1]
            year_months_map[year].append(ml)

        sorted_years  = sorted(year_months_map.keys())
        year_stat_cols = []
        for year in sorted_years:
            ycols = year_months_map[year]
            result[f'Tổng {year}']  = result[ycols].sum(axis=1)
            result[f'TB {year}']    = (result[f'Tổng {year}'] / len(ycols)).round(0)
            year_stat_cols.extend([f'Tổng {year}', f'TB {year}'])

        result['Tổng Toàn TG'] = md.sum(axis=1)

        last6 = mlabels[-6:] if len(mlabels) >= 6 else mlabels
        result['Tổng 6T Gần Nhất'] = result[last6].sum(axis=1)
        result['TB Tháng (6T GN)'] = (result['Tổng 6T Gần Nhất'] / len(last6)).round(0)

        # Doanh thu tổng
        if SALES['revenue'] in sales.columns:
            sales[SALES['revenue']] = pd.to_numeric(sales[SALES['revenue']], errors='coerce').fillna(0)
            rev_agg = sales.groupby(SALES['sku'])[SALES['revenue']].sum().reset_index()
            rev_agg.columns = ['SKU', 'Doanh Thu Tổng']
            rev_agg['SKU'] = rev_agg['SKU'].astype(str).str.strip()
            result = result.merge(rev_agg, on='SKU', how='left')
            result['Doanh Thu Tổng'] = result['Doanh Thu Tổng'].fillna(0)
            has_revenue = True
        else:
            has_revenue = False

        # Tổng bán theo Sale Team
        team_labels = []
        if SALES['sale_team'] in sales.columns:
            team_agg = (
                sales.groupby([SALES['sku'], SALES['sale_team']])[SALES['qty']]
                .sum().reset_index()
            )
            team_agg['SKU'] = team_agg[SALES['sku']].astype(str).str.strip()
            teams = sorted(team_agg[SALES['sale_team']].dropna().unique())

            team_wide = team_agg.pivot_table(
                index='SKU', columns=SALES['sale_team'],
                values=SALES['qty'], fill_value=0
            ).reset_index()
            team_wide.columns.name = None
            team_labels = [c for c in team_wide.columns if c != 'SKU']

            result = result.merge(team_wide, on='SKU', how='left')
            for tl in team_labels:
                result[tl] = result[tl].fillna(0)

            self.log(f'   Sale Team: {", ".join(str(t) for t in teams)}')

        result.attrs['year_stat_cols'] = year_stat_cols
        result.attrs['team_labels']    = team_labels
        result.attrs['has_revenue']    = has_revenue

        # 6. Tồn kho
        self.log('📦  Đọc file tồn kho…')
        inv = pd.read_excel(inv_path)
        inv.columns = [str(c).strip() for c in inv.columns]
        self._check_cols(inv, [INV['sku'], INV['qty']], 'file tồn kho')
        inv[INV['qty']] = pd.to_numeric(inv[INV['qty']], errors='coerce').fillna(0)

        has_reserved = INV['qty_reserved'] in inv.columns
        if has_reserved:
            inv[INV['qty_reserved']] = pd.to_numeric(inv[INV['qty_reserved']], errors='coerce').fillna(0)

        inv_f = inv.copy()
        if sel_brands and INV['brand'] in inv_f.columns:
            inv_f = inv_f[inv_f[INV['brand']].astype(str).str.strip().isin(sel_brands)]

        # Aggregate: Số lượng + Số lượng bảo lưu
        agg_cols = {INV['qty']: 'sum'}
        if has_reserved:
            agg_cols[INV['qty_reserved']] = 'sum'

        inv_agg = inv_f.groupby(INV['sku']).agg(agg_cols).reset_index()
        inv_agg.columns = ['SKU', 'Tồn Kho (Số Lượng)'] + (['Tồn Bảo Lưu'] if has_reserved else [])
        inv_agg['SKU'] = inv_agg['SKU'].astype(str).str.strip()

        if has_reserved:
            inv_agg['Tồn Khả Dụng'] = (inv_agg['Tồn Kho (Số Lượng)'] - inv_agg['Tồn Bảo Lưu']).clip(lower=0)
        else:
            inv_agg['Tồn Khả Dụng'] = inv_agg['Tồn Kho (Số Lượng)']

        self.log(f'   {inv_agg["SKU"].nunique():,} SKU  ·  Tổng tồn: {inv_agg["Tồn Kho (Số Lượng)"].sum():,.0f}  ·  Bảo lưu: {inv_agg["Tồn Bảo Lưu"].sum():,.0f}' if has_reserved else
                 f'   {inv_agg["SKU"].nunique():,} SKU  ·  Tổng tồn: {inv_agg["Tồn Kho (Số Lượng)"].sum():,.0f}')

        # Thông tin Brand/Category từ tồn kho (dùng cho SKU chưa có lịch sử bán)
        inv_info_map = {INV['sku']: 'SKU'}
        if INV['brand'] in inv_f.columns:
            inv_info_map[INV['brand']] = 'Brand'
        if INV['category'] in inv_f.columns:
            inv_info_map[INV['category']] = 'Category'
        inv_info = (inv_f[list(inv_info_map.keys())]
                    .rename(columns=inv_info_map).copy())
        inv_info['SKU'] = inv_info['SKU'].astype(str).str.strip()
        inv_info = inv_info.drop_duplicates(subset=['SKU'], keep='first')

        had_sales_skus = set(prod_info['SKU'].tolist())

        # Outer merge để giữ lại SKU tồn kho chưa có lịch sử bán
        result = result.merge(inv_agg, on='SKU', how='outer')

        # Điền Brand/Category cho SKU chỉ có trong tồn kho
        if len(inv_info.columns) > 1:
            result = result.merge(inv_info, on='SKU', how='left', suffixes=('', '_inv'))
            for col in ['Brand', 'Category']:
                if f'{col}_inv' in result.columns:
                    result[col] = result[col].fillna(result[f'{col}_inv'])
                    result.drop(columns=[f'{col}_inv'], inplace=True)

        # Điền 0 cho cột số bị thiếu (SKU chỉ có trong tồn kho)
        for col in ['Tồn Kho (Số Lượng)', 'Tồn Bảo Lưu', 'Tồn Khả Dụng']:
            if col in result.columns:
                result[col] = result[col].fillna(0)
        for col in mlabels + year_stat_cols + ['Tổng Toàn TG', 'Tổng 6T Gần Nhất', 'TB Tháng (6T GN)']:
            if col in result.columns:
                result[col] = result[col].fillna(0)

        # 7. Gợi ý đặt hàng — luôn tính Tổng + BL + BS
        total_months_val = months + leadtime
        scol = f'Gợi Ý Đặt Hàng (Tồn {months}T + LT {leadtime}T)'
        result[scol] = (
            (result['TB Tháng (6T GN)'] * total_months_val) - result['Tồn Khả Dụng']
        ).clip(lower=0).round(0)

        # 7b. Tính TB & Gợi Ý theo kênh BL / BS
        tb_bl_col = gy_bl_col = tb_bs_col = gy_bs_col = None

        def _channel_suggest(sales_df, team_filter, chan_name):
            if SALES['sale_team'] not in sales_df.columns:
                return None, None, None
            ch = (sales_df[sales_df[SALES['sale_team']].astype(str).isin(team_filter)]
                  if team_filter else sales_df)
            if ch.empty:
                return None, None, None
            ch_mo = (ch.groupby([SALES['sku'], '_month'])[SALES['qty']]
                     .sum().reset_index())
            ch_mo.columns = ['SKU', 'Month', 'Qty']
            ch_mo['SKU'] = ch_mo['SKU'].astype(str).str.strip()
            ch_piv = ch_mo.pivot_table(index='SKU', columns='Month',
                                       values='Qty', fill_value=0).reset_index()
            ch_piv.rename(columns=mcmap, inplace=True)
            last6_ch = [c for c in last6 if c in ch_piv.columns]
            n6 = len(last6_ch) if last6_ch else max(len(last6), 1)
            ch_piv['_tb'] = (
                ch_piv[last6_ch].sum(axis=1) / n6 if last6_ch else 0
            ).round(0)
            tb_col = f'TB {chan_name} (6T GN)'
            gy_col = f'Gợi Ý {chan_name} (Tồn {months}T + LT {leadtime}T)'
            tb_map = ch_piv.set_index('SKU')['_tb'].to_dict()
            return tb_col, gy_col, tb_map

        tb_bl_col, gy_bl_col, tb_bl_map = _channel_suggest(sales, bl_teams, 'BL')
        if tb_bl_col:
            result[tb_bl_col] = result['SKU'].map(tb_bl_map).fillna(0).round(0)
            result[gy_bl_col] = (
                result[tb_bl_col] * total_months_val - result['Tồn Khả Dụng']
            ).clip(lower=0).round(0)

        tb_bs_col, gy_bs_col, tb_bs_map = _channel_suggest(sales, bs_teams, 'BS')
        if tb_bs_col:
            result[tb_bs_col] = result['SKU'].map(tb_bs_map).fillna(0).round(0)
            result[gy_bs_col] = (
                result[tb_bs_col] * total_months_val - result['Tồn Khả Dụng']
            ).clip(lower=0).round(0)

        # 8. Nhận Xét (dựa trên Gợi Ý Tổng)
        nxcol = 'Nhận Xét'
        result[nxcol] = 'Đủ Hàng'
        result.loc[result['SKU'].isin(had_sales_skus) & (result[scol] >= 1), nxcol] = 'Cần Đặt'
        result.loc[~result['SKU'].isin(had_sales_skus), nxcol] = 'Chưa Có Lịch Sử Bán'

        # 9. Đặt Thực (để trống, người dùng tự nhập)
        dtcol = 'Đặt Thực'
        result[dtcol] = ''

        # Sắp xếp lại thứ tự cột
        inv_out        = [c for c in ['Tồn Kho (Số Lượng)', 'Tồn Bảo Lưu', 'Tồn Khả Dụng'] if c in result.columns]
        rev_out        = ['Doanh Thu Tổng'] if has_revenue else []
        extra_info_out = [c for c in extra_names if c in result.columns]
        tb_chan_out    = [c for c in [tb_bl_col, tb_bs_col] if c]
        gy_chan_out    = [c for c in [gy_bl_col, gy_bs_col] if c]
        ordered = (['SKU', 'Tên Sản Phẩm', 'Brand', 'Category'] + extra_info_out
                   + mlabels + year_stat_cols + ['Tổng Toàn TG']
                   + rev_out + team_labels
                   + ['Tổng 6T Gần Nhất', 'TB Tháng (6T GN)'] + tb_chan_out
                   + inv_out + [scol] + gy_chan_out + [dtcol, nxcol])
        result = result[[c for c in ordered if c in result.columns]]

        # Metadata cho _export
        result.attrs['scol']      = scol
        result.attrs['gy_bl_col'] = gy_bl_col
        result.attrs['gy_bs_col'] = gy_bs_col
        result.attrs['tb_bl_col'] = tb_bl_col
        result.attrs['tb_bs_col'] = tb_bs_col
        _text_fixed = ['Tên Sản Phẩm', 'Brand', 'Category', 'Model', 'Color', 'Frame Size', 'Sub Category']
        text_cols = [c for c in result.columns if c in _text_fixed]
        num_cols  = [c for c in result.columns if c not in text_cols + ['SKU', nxcol, dtcol]]
        result[text_cols] = result[text_cols].fillna('')
        result[num_cols]  = result[num_cols].fillna(0)
        no_hist = (~result['SKU'].isin(had_sales_skus)).sum()
        self.log(f'✅  {len(result):,} SKU  ·  {no_hist:,} SKU chưa có lịch sử bán')
        return result

    def _check_cols(self, df, required, label):
        miss = [c for c in required if c not in df.columns]
        if miss:
            raise ValueError(f'Thiếu cột trong {label}: {miss}\nCột hiện có: {list(df.columns)}')

    # ── Export Excel ──────────────────────────────────────────────────────────

    def _export(self, df: pd.DataFrame, months: int, leadtime: int,
                sel_brands: list, sel_cats: list) -> str:
        self.log('📝  Xuất Excel…')

        desktop = os.path.join(os.path.expanduser('~'), 'Desktop')
        ts      = datetime.now().strftime('%Y%m%d_%H%M%S')

        tag_parts = []
        if sel_brands:
            tag_parts.append('_'.join(sel_brands[:2]) + (f'+{len(sel_brands)-2}' if len(sel_brands) > 2 else ''))
        if sel_cats:
            tag_parts.append('_'.join(sel_cats[:2]) + (f'+{len(sel_cats)-2}' if len(sel_cats) > 2 else ''))
        tag = ('_' + '_'.join(tag_parts)) if tag_parts else '_TatCa'
        out = os.path.join(desktop, f'GoiYDatHang_{months}T{tag}_{ts}.xlsx')

        # Sort theo Doanh Thu Tổng giảm dần trước khi xuất
        if 'Doanh Thu Tổng' in df.columns:
            df = df.sort_values('Doanh Thu Tổng', ascending=False).reset_index(drop=True)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f'Gợi Ý {months}T'

        cols       = list(df.columns)
        scol       = df.attrs.get('scol', f'Gợi Ý Đặt Hàng (Tồn {months}T + LT {leadtime}T)')
        gy_bl_col  = df.attrs.get('gy_bl_col')
        gy_bs_col  = df.attrs.get('gy_bs_col')
        tb_bl_col  = df.attrs.get('tb_bl_col')
        tb_bs_col  = df.attrs.get('tb_bs_col')
        tb_chan_cols = [c for c in [tb_bl_col, tb_bs_col] if c and c in cols]
        dtcol      = 'Đặt Thực'
        fixed      = ['SKU', 'Tên Sản Phẩm', 'Brand', 'Category', 'Model',
                      'Color', 'Frame Size', 'Sub Category']
        month_cols = [c for c in cols if re.match(r'^Th\.\d{1,2}/\d{4}$', c)]
        inv_cols   = ['Tồn Kho (Số Lượng)', 'Tồn Bảo Lưu', 'Tồn Khả Dụng']
        stat6_cols = ['Tổng 6T Gần Nhất', 'TB Tháng (6T GN)']
        nxcol      = 'Nhận Xét'
        # Cột không nên SUBTOTAL (text hoặc trung bình)
        NOT_SUM    = set(fixed) | {c for c in cols if c.startswith('TB ')} | {nxcol, dtcol}

        # Detect year stat columns  (Tổng YYYY / TB YYYY / Tổng Toàn TG)
        year_total_cols = [c for c in cols if re.match(r'^Tổng \d{4}$', c)]
        year_avg_cols   = [c for c in cols if re.match(r'^TB \d{4}$', c)]
        grand_col       = 'Tổng Toàn TG'
        team_cols       = df.attrs.get('team_labels', [])

        # ── Màu header theo nhóm ──────────────────────────────────────────────
        # (bg_header, fg_header, fill_odd, fill_even)
        GRP_STYLE = {
            'fixed':      ('1A56DB', 'FFFFFF', 'FFFFFF',  'F0F4FF'),
            'month':      ('2563EB', 'FFFFFF', 'EFF6FF',  'DBEAFE'),
            'yr_total':   ('6D28D9', 'FFFFFF', 'F5F3FF',  'EDE9FE'),
            'yr_avg':     ('7C3AED', 'FFFFFF', 'FAF5FF',  'F3E8FF'),
            'grand':      ('4C1D95', 'FFFFFF', 'EDE9FE',  'DDD6FE'),
            'revenue':    ('B45309', 'FFFFFF', 'FFFBEB',  'FDE68A'),
            'team':       ('0F766E', 'FFFFFF', 'F0FDFA',  'CCFBF1'),
            'stat6':      ('0369A1', 'FFFFFF', 'E0F2FE',  'BAE6FD'),
            'tb_chan':    ('0369A1', 'FFFFFF', 'E0F2FE',  'BAE6FD'),
            'inv':        ('B45309', 'FFFFFF', 'FFFBEB',  'FEF3C7'),
            'suggest':    ('065F46', 'FFFFFF', 'ECFDF5',  'D1FAE5'),
            'suggest_bl': ('1D4ED8', 'FFFFFF', 'EFF6FF',  'BFDBFE'),
            'suggest_bs': ('6D28D9', 'FFFFFF', 'F5F3FF',  'DDD6FE'),
            'dat_thuc':   ('B45309', 'FFFFFF', 'FFFBEB',  'FEF9C3'),
            'nhan_xet':   ('374151', 'FFFFFF', 'F9FAFB',  'F3F4F6'),
        }

        def grp(c):
            if c in fixed:                       return 'fixed'
            if c in month_cols:                  return 'month'
            if c in year_total_cols:             return 'yr_total'
            if c in year_avg_cols:               return 'yr_avg'
            if c == grand_col:                   return 'grand'
            if c == 'Doanh Thu Tổng':            return 'revenue'
            if c in team_cols:                   return 'team'
            if c in stat6_cols:                  return 'stat6'
            if c in tb_chan_cols:                return 'tb_chan'
            if c in inv_cols:                    return 'inv'
            if c == scol:                        return 'suggest'
            if gy_bl_col and c == gy_bl_col:     return 'suggest_bl'
            if gy_bs_col and c == gy_bs_col:     return 'suggest_bs'
            if c == dtcol:                       return 'dat_thuc'
            if c == nxcol:                       return 'nhan_xet'
            return 'fixed'

        thin = Border(
            left=Side(style='thin',  color='D1D5DB'),
            right=Side(style='thin', color='D1D5DB'),
            top=Side(style='thin',   color='D1D5DB'),
            bottom=Side(style='thin',color='D1D5DB'),
        )
        ctr = Alignment(horizontal='center', vertical='center', wrap_text=True)
        lft = Alignment(horizontal='left',   vertical='center')
        rgt = Alignment(horizontal='right',  vertical='center')

        # ── Row 1: Subtotal ───────────────────────────────────────────────────
        for ci, col in enumerate(cols, 1):
            st        = GRP_STYLE[grp(col)]
            cell      = ws.cell(row=1, column=ci)
            col_ltr   = get_column_letter(ci)
            cell.fill   = PatternFill('solid', fgColor=st[0])
            cell.border = thin

            if col in NOT_SUM:
                cell.value     = 'SUBTOTAL' if ci == 1 else None
                cell.font      = Font(name='Calibri', bold=True, color=st[1], size=10)
                cell.alignment = ctr
            else:
                cell.value     = f'=SUBTOTAL(9,{col_ltr}3:{col_ltr}1048576)'
                cell.font      = Font(name='Calibri', bold=True, color=st[1], size=10)
                cell.alignment = rgt
                if col == 'Doanh Thu Tổng':
                    cell.number_format = '#,##0'

        ws.row_dimensions[1].height = 26

        # ── Row 2: Header ─────────────────────────────────────────────────────
        for ci, col in enumerate(cols, 1):
            st = GRP_STYLE[grp(col)]
            cell = ws.cell(row=2, column=ci, value=col)
            cell.font      = Font(name='Calibri', bold=True, color=st[1], size=10)
            cell.fill      = PatternFill('solid', fgColor=st[0])
            cell.alignment = ctr
            cell.border    = thin
        ws.row_dimensions[2].height = 38

        # ── Chuẩn bị công thức Excel ─────────────────────────────────────────
        col_letter   = {col: get_column_letter(ci) for ci, col in enumerate(cols, 1)}
        year_mcols: dict = {}
        for mc in month_cols:
            yr = mc.split('/')[1]
            year_mcols.setdefault(yr, []).append(mc)
        last6_cols   = month_cols[-6:] if len(month_cols) >= 6 else month_cols
        last6_n      = len(last6_cols) if last6_cols else 1
        total_months = months + leadtime

        def excel_formula(col, ri):
            cl = col_letter.get
            # Tồn Khả Dụng = MAX(0, Tồn Kho − Tồn Bảo Lưu)
            if col == 'Tồn Khả Dụng':
                tk = cl('Tồn Kho (Số Lượng)')
                if not tk: return None
                if 'Tồn Bảo Lưu' in col_letter:
                    return f'=MAX(0,{tk}{ri}-{cl("Tồn Bảo Lưu")}{ri})'
                return f'={tk}{ri}'
            # Tổng YYYY / TB YYYY
            for yr, ycols in year_mcols.items():
                if col == f'Tổng {yr}':
                    return f'=SUM({cl(ycols[0])}{ri}:{cl(ycols[-1])}{ri})'
                if col == f'TB {yr}':
                    return f'=ROUND({cl(f"Tổng {yr}")}{ri}/{len(ycols)},0)'
            # Tổng Toàn TG = SUM tất cả tháng
            if col == 'Tổng Toàn TG' and month_cols:
                return f'=SUM({cl(month_cols[0])}{ri}:{cl(month_cols[-1])}{ri})'
            # Tổng 6T Gần Nhất
            if col == 'Tổng 6T Gần Nhất' and last6_cols:
                return f'=SUM({cl(last6_cols[0])}{ri}:{cl(last6_cols[-1])}{ri})'
            # TB Tháng (6T GN) = Tổng 6T / n
            if col == 'TB Tháng (6T GN)' and 'Tổng 6T Gần Nhất' in col_letter:
                return f'=ROUND({cl("Tổng 6T Gần Nhất")}{ri}/{last6_n},0)'
            # Gợi Ý Tổng
            if col == scol:
                tb = cl('TB Tháng (6T GN)'); tkd = cl('Tồn Khả Dụng')
                if tb and tkd:
                    return f'=MAX(0,{tb}{ri}*{total_months}-{tkd}{ri})'
            # Gợi Ý BL
            if gy_bl_col and col == gy_bl_col and tb_bl_col:
                tb = cl(tb_bl_col); tkd = cl('Tồn Khả Dụng')
                if tb and tkd:
                    return f'=MAX(0,{tb}{ri}*{total_months}-{tkd}{ri})'
            # Gợi Ý BS
            if gy_bs_col and col == gy_bs_col and tb_bs_col:
                tb = cl(tb_bs_col); tkd = cl('Tồn Khả Dụng')
                if tb and tkd:
                    return f'=MAX(0,{tb}{ri}*{total_months}-{tkd}{ri})'
            # Nhận Xét (công thức IF lồng nhau)
            if col == nxcol:
                tb = cl('TB Tháng (6T GN)'); ttg = cl('Tổng Toàn TG'); gy = cl(scol)
                if tb and ttg and gy:
                    return (f'=IF(AND({tb}{ri}=0,{ttg}{ri}=0),"Chưa Có Lịch Sử Bán",'
                            f'IF({gy}{ri}>=1,"Cần Đặt","Đủ Hàng"))')
            return None

        # ── Data rows với màu xen kẽ (bắt đầu từ row 3) ─────────────────────
        for ri, (_, row) in enumerate(df.iterrows(), start=3):
            is_odd = (ri % 2 == 0)   # hàng chẵn (dữ liệu lẻ) = nền nhạt hơn
            for ci, col in enumerate(cols, 1):
                val = row[col]
                if pd.isna(val):     val = 0
                elif isinstance(val, float) and val.is_integer(): val = int(val)

                formula = excel_formula(col, ri)
                cell    = ws.cell(row=ri, column=ci,
                                  value=formula if formula is not None else val)
                cell.border = thin
                g  = grp(col)
                st = GRP_STYLE[g]
                # Suggest giữ màu riêng; dat_thuc/nhan_xet tự override bên dưới
                if g in ('suggest', 'suggest_bl', 'suggest_bs', 'tb_chan'):
                    fill_color = st[2] if is_odd else st[3]
                else:
                    fill_color = 'FFFFFF' if is_odd else 'F0F4FF'
                cell.fill   = PatternFill('solid', fgColor=fill_color)

                if g == 'fixed':
                    cell.font      = Font(name='Calibri', size=10)
                    cell.alignment = ctr if col == 'SKU' else lft
                elif g == 'month':
                    cell.font      = Font(name='Calibri', size=10)
                    cell.alignment = ctr
                elif g in ('yr_total', 'grand', 'team'):
                    cell.font      = Font(name='Calibri', bold=True, size=10)
                    cell.alignment = rgt
                elif g == 'revenue':
                    cell.font         = Font(name='Calibri', bold=True, size=10)
                    cell.alignment    = rgt
                    cell.number_format = '#,##0'
                elif g == 'yr_avg':
                    cell.font          = Font(name='Calibri', size=10, italic=True)
                    cell.alignment     = rgt
                    cell.number_format = '0'
                elif g == 'stat6':
                    cell.font          = Font(name='Calibri', bold=True, size=10)
                    cell.alignment     = rgt
                    if col == 'TB Tháng (6T GN)':
                        cell.number_format = '0'
                elif g == 'inv':
                    cell.font      = Font(name='Calibri', bold=True, size=10)
                    cell.alignment = ctr
                elif g in ('suggest', 'suggest_bl', 'suggest_bs'):
                    cell.font      = Font(name='Calibri', bold=True, size=11)
                    cell.alignment = ctr
                elif g == 'tb_chan':
                    cell.font      = Font(name='Calibri', bold=True, size=10)
                    cell.alignment = rgt
                elif g == 'dat_thuc':
                    cell.fill      = PatternFill('solid', fgColor='FFFBEB')
                    cell.font      = Font(name='Calibri', bold=True, size=11)
                    cell.alignment = ctr
                    cell.value     = None   # luôn để trống cho người dùng nhập
                elif g == 'nhan_xet':
                    cell.alignment = ctr
                    if val == 'Cần Đặt':
                        cell.fill = PatternFill('solid', fgColor='FEF3C7')
                        cell.font = Font(name='Calibri', bold=True, size=10, color='92400E')
                    elif val == 'Chưa Có Lịch Sử Bán':
                        cell.fill = PatternFill('solid', fgColor='F3F4F6')
                        cell.font = Font(name='Calibri', size=10, italic=True, color='6B7280')
                    else:  # Đủ Hàng
                        cell.fill = PatternFill('solid', fgColor='ECFDF5')
                        cell.font = Font(name='Calibri', size=10, color='065F46')
            ws.row_dimensions[ri].height = 18

        # ── Độ rộng cột ───────────────────────────────────────────────────────
        W = {
            'SKU': 18, 'Tên Sản Phẩm': 42, 'Brand': 14, 'Category': 18,
            'Color': 14, 'Frame Size': 13, 'Sub Category': 18,
            grand_col: 16, 'Tổng 6T Gần Nhất': 16, 'TB Tháng (6T GN)': 16,
            'Tồn Kho (Số Lượng)': 17, 'Tồn Bảo Lưu': 13, 'Tồn Khả Dụng': 14,
            scol: 26, dtcol: 12, nxcol: 22,
        }
        if tb_bl_col: W[tb_bl_col] = 16
        if tb_bs_col: W[tb_bs_col] = 16
        if gy_bl_col: W[gy_bl_col] = 26
        if gy_bs_col: W[gy_bs_col] = 26
        for c in year_total_cols + year_avg_cols:
            W[c] = 13
        for c in team_cols:
            W[c] = 16
        W['Doanh Thu Tổng'] = 18

        for ci, col in enumerate(cols, 1):
            ws.column_dimensions[get_column_letter(ci)].width = W.get(
                col, 10 if col in month_cols else 14)

        # ── Nhóm cột tháng theo năm (outline, mặc định thu gọn) ─────────────────
        year_month_cidxs: dict[str, list] = {}
        for ci, col in enumerate(cols, 1):
            m = re.match(r'^Th\.\d{1,2}/(\d{4})$', col)
            if m:
                year_month_cidxs.setdefault(m.group(1), []).append(ci)

        for year in sorted(year_month_cidxs):
            for ci in year_month_cidxs[year]:
                ltr = get_column_letter(ci)
                ws.column_dimensions[ltr].outline_level = 1
                ws.column_dimensions[ltr].hidden = True

        if year_month_cidxs:
            ws.sheet_format.outlineLevelCol = 1

        # ── Freeze + filter ───────────────────────────────────────────────────
        ws.freeze_panes = 'C3'   # cố định 2 dòng đầu + cột A, B
        ws.auto_filter.ref = f'A2:{get_column_letter(len(cols))}2'

        # ── Tab màu ───────────────────────────────────────────────────────────
        ws.sheet_properties.tabColor = '0071E3'

        # ── Sheet 2: Hướng Dẫn & Logic Tính Toán ─────────────────────────────
        ws2 = wb.create_sheet(title='Hướng Dẫn & Logic')
        ws2.sheet_properties.tabColor = '6D28D9'
        ws2.column_dimensions['A'].width = 28
        ws2.column_dimensions['B'].width = 22
        ws2.column_dimensions['C'].width = 62
        ws2.column_dimensions['D'].width = 42

        def _cell(r, c, val, bold=False, bg=None, fg='000000', size=10,
                  italic=False, wrap=False, align='left'):
            cell = ws2.cell(row=r, column=c, value=val)
            cell.font = Font(name='Calibri', bold=bold, italic=italic,
                             color=fg, size=size)
            if bg:
                cell.fill = PatternFill('solid', fgColor=bg)
            cell.alignment = Alignment(
                horizontal=align, vertical='center',
                wrap_text=wrap)
            cell.border = Border(
                left=Side(style='thin', color='E5E7EB'),
                right=Side(style='thin', color='E5E7EB'),
                top=Side(style='thin', color='E5E7EB'),
                bottom=Side(style='thin', color='E5E7EB'),
            )
            return cell

        def _section(r, title):
            """Dòng tiêu đề section."""
            for c in range(1, 5):
                ws2.cell(row=r, column=c).fill = PatternFill('solid', fgColor='1C1C1E')
            cell = ws2.cell(row=r, column=1, value=title)
            cell.font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
            cell.alignment = Alignment(horizontal='left', vertical='center')
            ws2.merge_cells(start_row=r, start_column=1,
                            end_row=r, end_column=4)
            ws2.row_dimensions[r].height = 22
            return r + 1

        def _header_row(r):
            labels = ['Tên Cột / Trường', 'Kiểu Dữ Liệu', 'Công Thức / Nguồn Dữ Liệu', 'Giải Thích']
            for ci, lbl in enumerate(labels, 1):
                _cell(r, ci, lbl, bold=True, bg='1A56DB', fg='FFFFFF', align='center')
            ws2.row_dimensions[r].height = 20
            return r + 1

        def _row(r, name, dtype, formula, explain, alt=False):
            bg = 'F8F9FF' if alt else 'FFFFFF'
            _cell(r, 1, name,    bold=True,  bg=bg, fg='1A56DB')
            _cell(r, 2, dtype,   bold=False, bg=bg, fg='374151', align='center')
            _cell(r, 3, formula, bold=False, bg=bg, fg='065F46', wrap=True)
            _cell(r, 4, explain, bold=False, bg=bg, fg='374151', wrap=True)
            ws2.row_dimensions[r].height = 42
            return r + 1

        run_ts   = datetime.now().strftime('%d/%m/%Y %H:%M')
        brands_s = ', '.join(sel_brands) if sel_brands else 'Tất cả'
        cats_s   = ', '.join(sel_cats)   if sel_cats   else 'Tất cả'

        # ── Tiêu đề file ──────────────────────────────────────────────────────
        ws2.merge_cells('A1:D1')
        t = ws2.cell(row=1, column=1,
                     value='GỢI Ý ĐẶT HÀNG — HƯỚNG DẪN & LOGIC TÍNH TOÁN')
        t.font      = Font(name='Calibri', bold=True, color='FFFFFF', size=14)
        t.fill      = PatternFill('solid', fgColor='1C1C1E')
        t.alignment = Alignment(horizontal='center', vertical='center')
        ws2.row_dimensions[1].height = 32

        # ── Thông số chạy ─────────────────────────────────────────────────────
        r = 2
        for label, val in [
            ('Thời gian xuất',        run_ts),
            ('Tồn kho tối thiểu',     f'{months} tháng'),
            ('Leadtime về hàng',      f'{leadtime} tháng'),
            ('Thương hiệu lọc',       brands_s),
            ('Danh mục lọc',          cats_s),
            ('Số tháng đặt hàng',     f'{months + leadtime} tháng (= Tồn tối thiểu + Leadtime)'),
        ]:
            ws2.merge_cells(start_row=r, start_column=3,
                            end_row=r, end_column=4)
            _cell(r, 1, label, bold=True,  bg='F3F4F6', fg='374151')
            _cell(r, 2, '',               bg='F3F4F6')
            _cell(r, 3, val,   bold=False, bg='FAFAFA', fg='000000', wrap=True)
            ws2.row_dimensions[r].height = 18
            r += 1

        r += 1  # khoảng trắng

        # ══════════════════════════════════════════════════════════════════════
        # SECTION 1 — Nguồn dữ liệu đầu vào
        # ══════════════════════════════════════════════════════════════════════
        r = _section(r, '  PHẦN 1 — NGUỒN DỮ LIỆU ĐẦU VÀO')
        r = _header_row(r)
        rows1 = [
            ('SKU',             'Text',   'File bán hàng · cột "SKU"',
             'Mã sản phẩm nội bộ — dùng làm khoá join giữa 2 file.'),
            ('Tên Sản Phẩm',    'Text',   'File bán hàng · cột "Product Item" (đã bỏ tiền tố [...])',
             'Tên hiển thị của sản phẩm.'),
            ('Brand',           'Text',   'File bán hàng · cột "Brand"',
             'Thương hiệu sản phẩm.'),
            ('Category',        'Text',   'File bán hàng · cột "Category"  |  File tồn kho · cột "Sản phẩm/Nhóm Điểm bán lẻ/Tên hiển thị"',
             'Danh mục sản phẩm. SKU chỉ có trong tồn kho thì lấy từ file tồn kho.'),
            ('Model / Color / Frame Size / Sub Category',
             'Text',
             'File bán hàng · cột "Model", "Color", "Frame Size", "Sub Category"',
             'Thông tin mô tả sản phẩm. Chỉ có ở sản phẩm đã từng bán.'),
            ('Tồn Kho (Số Lượng)', 'Số',  'File tồn kho · SUM(Số lượng) gom theo SKU',
             'Tổng số lượng tồn kho thực tế trên toàn bộ địa điểm.'),
            ('Tồn Bảo Lưu',     'Số',     'File tồn kho · SUM(Số lượng bảo lưu) gom theo SKU',
             'Số lượng đã được giữ chỗ / đặt cọc, chưa xuất nhưng không còn trống.'),
            ('Tồn Khả Dụng',    'Số',     'Tồn Kho (Số Lượng)  −  Tồn Bảo Lưu  ≥ 0',
             'Số lượng thực sự có thể bán hoặc điều phối. Luôn ≥ 0.'),
        ]
        for i, args in enumerate(rows1):
            r = _row(r, *args, alt=(i % 2 == 1))

        r += 1

        # ══════════════════════════════════════════════════════════════════════
        # SECTION 2 — Số liệu bán hàng
        # ══════════════════════════════════════════════════════════════════════
        r = _section(r, '  PHẦN 2 — SỐ LIỆU BÁN HÀNG THEO THÁNG')
        r = _header_row(r)
        rows2 = [
            ('Th.M/YYYY',       'Số',
             'SUM(Quantity) của SKU trong tháng M năm YYYY\n(lọc bỏ SERVICE, COUPON, DISCOUNT)',
             'Số lượng bán ra trong từng tháng. Các tháng được nhóm theo năm và mặc định ẩn — bấm [+] trên header để mở.'),
            ('Tổng YYYY',       'Số',
             'SUM(Th.1/YYYY : Th.12/YYYY)  —  cộng tất cả tháng trong năm',
             'Tổng số lượng bán cả năm YYYY.'),
            ('TB YYYY',         'Số',
             'Tổng YYYY  ÷  số tháng có dữ liệu trong năm YYYY',
             'Trung bình bán mỗi tháng của năm YYYY.'),
            ('Tổng Toàn TG',    'Số',
             'SUM(tất cả tháng trong file)',
             'Tổng số lượng bán toàn bộ thời gian có dữ liệu.'),
            ('Doanh Thu Tổng',  'Số (VNĐ)',
             'SUM(Revenue) của SKU — toàn bộ thời gian',
             'Tổng doanh thu. File xuất được sắp xếp giảm dần theo cột này.'),
            ('Tổng 6T Gần Nhất','Số',
             'SUM(6 tháng gần nhất trong file)',
             'Tổng số lượng bán của 6 tháng gần nhất — dùng để tính TB tháng.'),
            ('TB Tháng (6T GN)', 'Số',
             'Tổng 6T Gần Nhất  ÷  6',
             '⭐ Sức bán bình quân tháng — CỘT CỐT LÕI để tính gợi ý đặt hàng.'),
        ]
        for i, args in enumerate(rows2):
            r = _row(r, *args, alt=(i % 2 == 1))

        r += 1

        # ══════════════════════════════════════════════════════════════════════
        # SECTION 3 — Công thức gợi ý
        # ══════════════════════════════════════════════════════════════════════
        r = _section(r, '  PHẦN 3 — CÔNG THỨC GỢI Ý ĐẶT HÀNG')
        r = _header_row(r)
        rows3 = [
            (scol,              'Số',
             f'MAX( 0,  TB Tháng (6T GN)  ×  ({months} + {leadtime})  −  Tồn Khả Dụng )\n'
             f'= MAX( 0,  TB Tháng  ×  {months + leadtime}  −  Tồn Khả Dụng )',
             f'Số lượng cần đặt thêm để đảm bảo:\n'
             f'  • Đủ hàng bán trong {months} tháng tới (tồn tối thiểu)\n'
             f'  • Đủ hàng bán trong {leadtime} tháng chờ hàng về (leadtime)\n'
             f'Nếu tồn hiện tại đã đủ → kết quả = 0.'),
            ('Đặt Thực',        'Số (nhập tay)',
             '(để trống — người dùng tự nhập)',
             'Cột nhập tay. Sau khi xem gợi ý, người dùng điền số lượng thực tế muốn đặt.'),
            ('Nhận Xét',        'Text',
             'Nếu SKU không có trong file bán hàng  →  "Chưa Có Lịch Sử Bán"\n'
             'Nếu Gợi Ý Đặt Hàng ≥ 1               →  "Cần Đặt"\n'
             'Nếu Gợi Ý Đặt Hàng = 0               →  "Đủ Hàng"',
             'Phân loại trạng thái để lọc nhanh:\n'
             '  🟡 Cần Đặt — thiếu hàng theo công thức\n'
             '  🟢 Đủ Hàng — tồn hiện tại đã đủ\n'
             '  ⚫ Chưa Có Lịch Sử Bán — SKU có tồn nhưng chưa bán lần nào'),
        ]
        for i, args in enumerate(rows3):
            r = _row(r, *args, alt=(i % 2 == 1))

        r += 1

        # ══════════════════════════════════════════════════════════════════════
        # SECTION 4 — Ví dụ minh hoạ
        # ══════════════════════════════════════════════════════════════════════
        r = _section(r, '  PHẦN 4 — VÍ DỤ MINH HOẠ')
        eg_bg  = 'FFFBEB'
        eg_bg2 = 'FEF3C7'

        eg_rows = [
            ('', 'TB Tháng (6T GN)', 'Tồn Tối Thiểu', 'Leadtime', 'Tồn Khả Dụng',
             'Gợi Ý = TB × (TT + LT) − Tồn', 'Nhận Xét'),
            ('SKU-A', '10',  f'{months}T', f'{leadtime}T', '5',
             f'10 × {months+leadtime} − 5 = {10*(months+leadtime)-5}', 'Cần Đặt' if 10*(months+leadtime)-5 > 0 else 'Đủ Hàng'),
            ('SKU-B', '10',  f'{months}T', f'{leadtime}T', f'{10*(months+leadtime)+10}',
             f'10 × {months+leadtime} − {10*(months+leadtime)+10} = 0 (âm → 0)', 'Đủ Hàng'),
            ('SKU-C', '0',   f'{months}T', f'{leadtime}T', '50',
             '0 × ... = 0', 'Chưa Có Lịch Sử Bán'),
        ]
        headers_eg = eg_rows[0]
        for ci, h in enumerate(headers_eg, 1):
            c2 = ws2.cell(row=r, column=ci, value=h)
            c2.font      = Font(name='Calibri', bold=True, color='92400E', size=10)
            c2.fill      = PatternFill('solid', fgColor=eg_bg2)
            c2.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            c2.border    = Border(
                left=Side(style='thin', color='E5E7EB'),
                right=Side(style='thin', color='E5E7EB'),
                top=Side(style='thin', color='E5E7EB'),
                bottom=Side(style='thin', color='E5E7EB'),
            )
        ws2.row_dimensions[r].height = 36
        ws2.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
        r += 1

        for data_row in eg_rows[1:]:
            for ci, v in enumerate(data_row, 1):
                c2 = ws2.cell(row=r, column=ci, value=v)
                c2.font      = Font(name='Calibri', size=10)
                c2.fill      = PatternFill('solid', fgColor=eg_bg)
                c2.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                c2.border    = Border(
                    left=Side(style='thin', color='E5E7EB'),
                    right=Side(style='thin', color='E5E7EB'),
                    top=Side(style='thin', color='E5E7EB'),
                    bottom=Side(style='thin', color='E5E7EB'),
                )
            ws2.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
            ws2.row_dimensions[r].height = 28
            r += 1

        # Độ rộng cột ví dụ
        for col_ltr, w in [('A',14),('B',18),('C',16),('D',14),('E',18),('F',38),('G',22)]:
            ws2.column_dimensions[col_ltr].width = w

        wb.save(out)
        self.log(f'   → {os.path.basename(out)}')
        return out


# ─── DPI-aware trên Windows (tránh bị mờ do bitmap scaling) ─────────────────

if platform.system() == 'Windows':
    try:
        import ctypes
        ctypes.windll.shcore.SetProcessDpiAwareness(2)   # Per-monitor DPI aware
    except Exception:
        try:
            ctypes.windll.shcore.SetProcessDpiAwareness(1)   # System DPI aware
        except Exception:
            try:
                ctypes.windll.user32.SetProcessDPIAware()    # Fallback Win 7/8
            except Exception:
                pass


# ─── Entry ───────────────────────────────────────────────────────────────────

def main():
    root = tk.Tk()
    App(root)
    root.update_idletasks()
    w, h = root.winfo_width(), root.winfo_height()
    x = (root.winfo_screenwidth()  - w) // 2
    y = (root.winfo_screenheight() - h) // 2
    root.geometry(f'+{x}+{y}')
    root.mainloop()


if __name__ == '__main__':
    main()
